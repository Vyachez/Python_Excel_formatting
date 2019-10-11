
# coding: utf-8

# # Magic Excel with Python by Vyachez

# ### used package https://openpyxl.readthedocs.io/en/stable/' 

# Openpyxl is a Python library to read/write Excel 2010 xlsx/xlsm/xltx/xltm files.

# importing modules
import os
import numpy as np
import pandas as pd

import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting import Rule
from openpyxl.styles import Font, PatternFill, Border, NamedStyle, Side, Alignment
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation


# #### Getting data
path = os.getcwd()+"/"
data_file = "test_data.csv"
dest_filename = "Spreadsheet_ready.xlsx"

data = pd.read_csv(path+data_file)
data.head(3)

# #### Initializing openpyxl and setting up data
# creating openpyxl object to read data in excel
wb = Workbook()

# defining tab
main_tab = wb.active
main_tab.title = "Main_tab"

# filling with data
for r in dataframe_to_rows(data, index=False, header=True):
    main_tab.append(r)
    
# Saving workbook
wb.save(filename = dest_filename)

# #### Formatting
# ##### defining formatting styles

# body style 
def add_body_style(wb):
    name = 'body'
    st = NamedStyle(name=name)
    st.font = Font(name='Calibri', bold=False, size=11)
    bd = Side(style='thin', color="000000")
    st.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    st.alignment=Alignment(horizontal='left',
                        vertical='center',
                        text_rotation=0,
                        wrap_text=True,
                        shrink_to_fit=False,
                        indent=0)
    st.fill = PatternFill(start_color='fefef5',
                       end_color='fefef5',
                       fill_type='solid')
    wb.add_named_style(st)
    return name

# header style
def add_head_style(wb):
    name = 'headstyle'
    st = NamedStyle(name=name)
    st.font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    bd = Side(style='thin', color="000000")
    st.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    st.alignment=Alignment(horizontal='center',
                        vertical='center',
                        text_rotation=0,
                        wrap_text=True,
                        shrink_to_fit=False,
                        indent=0)
    st.fill = PatternFill(start_color='538DD5',
                       end_color='538DD5',
                       fill_type='solid')
    wb.add_named_style(st)
    return name

# highlighter style 
def add_highlighter(wb):
    name = 'highlight'
    st = NamedStyle(name=name)
    st.font = Font(name='Calibri', bold=True, size=11)
    bdb = Side(style='medium', color="000000")
    bdt = Side(style='thin', color="000000")
    st.border = Border(left=bdt, top=bdt, right=bdt, bottom=bdt)
    st.alignment=Alignment(horizontal='left',
                        vertical='center',
                        text_rotation=0,
                        wrap_text=True,
                        shrink_to_fit=False,
                        indent=0)
    st.fill = PatternFill(start_color='ffcccc',
                       end_color='ffcccc',
                       fill_type='solid')
    wb.add_named_style(st)
    return name

# left column style
def add_leftcol_style(wb):
    name = 'indexer'
    st = NamedStyle(name=name)
    st.font = Font(name='Calibri', bold=True, size=10)
    bd = Side(style='thin', color="000000")
    st.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    st.alignment=Alignment(horizontal='left',
                        vertical='center',
                        text_rotation=0,
                        wrap_text=True,
                        shrink_to_fit=False,
                        indent=0)
    st.fill = PatternFill(start_color='e0e0eb',
                       end_color='e0e0eb',
                       fill_type='solid')
    wb.add_named_style(st)
    return name

# ##### applying styles 
# appending styles to workbook
letters = ["A","B","C","D","E","F","G"]

# left column style
index_style = add_leftcol_style(wb)
for rw in range(2, 14):
    main_tab['A'+str(rw)].style = index_style
    
# header style
head_style = add_head_style(wb)
for l in letters:
    main_tab[l+"1"].style = head_style

# body style
body_style = add_body_style(wb)
highlighter = add_highlighter(wb) # highlight critical risk levels
for rw in range(2, 14):
    for l in letters[1:]:
        main_tab[l+str(rw)].style = body_style
        if main_tab['F'+str(rw)].value < 2:
            main_tab[l+str(rw)].style = highlighter
            main_tab['A'+str(rw)].style = highlighter

# ##### rows and columns dimensions
# applying rows dimensions
# header height
main_tab.row_dimensions[1].height = 30

# regular row height
for dim in range(2, 14):
    main_tab.row_dimensions[dim].height = 40
    
# applying columns dimensions
# dimensiton for first columns (as an example)
main_tab.column_dimensions['A'].width = 17

# iterating through number of columns
for dim in range(1, 2+1):
    main_tab.column_dimensions[letters[dim]].width = 20
for dim in range(3, 5+1):
    main_tab.column_dimensions[letters[dim]].width = 12

# finishing with individual columns
main_tab.column_dimensions['G'].width = 25

# ##### conditional formatting
# adds conditional format to selected range
def add_cond_text_format(ws, text, color, start, end):
    '''
    Takes:
    - ws - worksheet object
    - text - as string
    - color - hex color
    - start cell+col string
    - end cell+col string
    '''
    fill = PatternFill(bgColor=color)
    dxf = DifferentialStyle(fill=fill)
    rule = Rule(type="cellIs", operator="equal", dxf=dxf)
    rule.formula = ['"{}"'.format(text)]
    ws.conditional_formatting.add(start+":"+end, rule)

# inserting conditional formatting formula for ratings
values = ['Red', 'Orange', 'Yellow', 'Green']
colors = ['FF0000','FFC000','FFFF00','92D050']
for val, colr in zip(values,colors):
    add_cond_text_format(main_tab, val, colr, 'E2', 'E13')

# ##### making filters
# filtering
main_tab.auto_filter.ref = "A1:G1"
# freezing
main_tab.freeze_panes = "A2"

# ##### other useful stuff
# hiding gridlines
main_tab.sheet_view.showGridLines = False

# putting reference at the end of document
main_tab['A15'].value = '* Just footnote.'

# view
main_tab.sheet_view.zoomScale = 110

# Saving workbook
wb.save(filename = dest_filename)
